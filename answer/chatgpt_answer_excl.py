# -*- coding: utf-8 -*-
from answer.chatgpt_answer import answer_bydoc
import openpyxl
import time


# -*- coding: utf-8 -*-
import xlsxwriter as xw


def xw_toExcel(data, fileName):  # xlsxwriter库储存数据到excel
    workbook = xw.Workbook(fileName)  # 创建工作簿
    worksheet1 = workbook.add_worksheet("sheet1")  # 创建子表
    worksheet1.activate()  # 激活表
    title = ['序号', '问题', '答案', '耗时']  # 设置表头
    worksheet1.write_row('A1', title)  # 从A1单元格开始写入表头
    i = 2  # 从第二行开始写入数据
    for j in range(len(data)):
        insertData = [data[j]["序号"], data[j]["问题"], data[j]["答案"], data[j]["耗时"]]
        row = 'A' + str(i)
        worksheet1.write_row(row, insertData)
        i += 1
    workbook.close()  # 关闭表


# excel表格转json文件
def xr_fromExcel(excel_file,sheet_no):
    # 加载工作薄
    wb = openpyxl.load_workbook(excel_file)
    # 获取sheet页
    sheet = wb.worksheets[sheet_no]
    max_row = sheet.max_row
    # 列数
    max_column = sheet.max_column
    print("max_row: %d, max_column: %d" % (max_row, max_column))
    # 结果，数组存储
    result = []
    heads = []
    # 解析表头
    for column in range(max_column):
        # 读取的话行列是从（1，1）开始
        heads.append(sheet.cell(1, column + 1).value)
    # 遍历每一行
    for row in range(max_row):
        if row == 0:
            continue
        one_line = {}
        cell = sheet.cell(row + 1 ,1)
        print("row:" + str(row) + "问题:" + str(cell.value))
        if cell.value == None:
            continue
        t1 = time.time()
        replay = answer_bydoc("my_doc1", cell.value)
        t2 = time.time()
        one_line['序号'] = row
        one_line['问题'] = cell.value
        one_line['答案'] = replay
        one_line['耗时'] = t2-t1
        print(one_line)
        result.append(one_line)
    new_file_url = str.replace(excel_file, ".xlsx", '_result.xlsx')
    xw_toExcel(result, new_file_url)
    wb.close()
    # 将json保存为文件

#main
if '__main__' == __name__:
     xr_fromExcel(u'/Users/miao/mydocs/个人/公司/邮乐/问题测试/zhang.xlsx', 0)
     #xr_fromExcel(u'/Users/miao/mydocs/个人/公司/邮乐/问题测试/test.xlsx',0)

