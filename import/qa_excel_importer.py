# -*- coding: utf-8 -*-
import openpyxl
from answer.chatgpt_answer import add_doc


# excel表格转json文件
def excel2milvus(excel_file, collection_name,sheet_no):
    # 加载工作薄
    wb = openpyxl.load_workbook(excel_file)
    # 获取sheet页
    sheet = wb.worksheets[sheet_no]
    max_row = sheet.max_row
    # 列数
    max_column = sheet.max_column
    print("max_row: %d, max_column: %d" % (max_row, max_column))
    # 结果，数组存储
    heads = []
    # 解析表头
    for column in range(max_column):
        # 读取的话行列是从（1，1）开始
        heads.append(sheet.cell(1, column + 1).value)
    # 遍历每一行
    for row in range(max_row):
        if row == 0:
            continue
        cell1 = sheet.cell(row + 1, 1)
        value1 = cell1.value
        cell2 = sheet.cell(row + 1, 2)
        value2 = cell2.value
        print(str(value2))
        add_doc(collection_name, str(value1)+"?\n", str(value2))

    wb.close()


#main
if '__main__' == __name__:
    #excel2milvus(u'/Users/miao/mydocs/个人/公司/邮乐/邮乐AI数字人知识库收集 (第2批次)(1226) 的副本.xlsx',"youle1",1)
    #excel2milvus(u'/Users/miao/mydocs/个人/公司/邮乐/张东旭.xlsx',"youle1",0)
    #excel2milvus(u'/Users/miao/mydocs/个人/公司/邮乐/1226补充信息.xlsx',"youle1",0)
    excel2milvus(u'/Users/miao/mydocs/个人/公司/邮乐/0102.xlsx',"youle1",0)


#excel2json(u'/Users/miao/mydocs/个人/公司/邮乐/资料库补充资料.xlsx', '/Users/miao/mydocs/个人/公司/邮乐1/3.doc',3)